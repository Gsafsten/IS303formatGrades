"""
Garrett Safsten, Jack Mair, Ryan Baldwin, Tanner Crookston
Description: Script to reorganize Excel data from an existing file with additional summary statistics, formatting, functions, and filters. 
"""
# This pulls in the libraries we will need.
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import statistics

# Load the original workbook
myWorkbook = openpyxl.load_workbook("Poorly_Organized_Data_1.xlsx")
currentSheet = myWorkbook.active

# Create a new workbook
newWorkbook = Workbook()
newWorkbook.remove(newWorkbook.active)  # Remove default sheet

# Dictionary to track the last row for each sheet
row_counters = {}

# Dictionary to store grades for each category
category_grades = {}

# Iterate through rows in the original data sheet
for row in currentSheet.iter_rows(min_row=2, values_only=True):
    category = row[0]  # Column A value (determines sheet name)
    full_name = row[1]  # Column B value (contains names)
    grade = row[2]  # Column C value (grade)
    
    # Make sure grade is a number
    if isinstance(grade, str):
        try:
            grade = float(grade)
        except ValueError:
            # If grade can't be converted to a number, skip this processing
            continue
    
    # Add grade to the category's list of grades
    if category not in category_grades:
        category_grades[category] = []
    category_grades[category].append(grade)
    
    # Ensure the sheet exists
    if category not in newWorkbook.sheetnames:
        newSheet = newWorkbook.create_sheet(title=str(category))
        newSheet.append(["Last Name", "First Name", "Student ID", "Grade"])  # Add headers
        row_counters[category] = 2  # Start row count at 2
        newSheet.auto_filter.ref = "A1:D1" # Sets a filter in the first row from column A to D
    
    # Split name into parts
    lstOrganizedData = full_name.split("_")
    
    # Store values in the appropriate sheet
    newSheet = newWorkbook[category]
    newSheet[f"A{row_counters[category]}"] = lstOrganizedData[0]
    newSheet[f"B{row_counters[category]}"] = lstOrganizedData[1]
    newSheet[f"C{row_counters[category]}"] = lstOrganizedData[2]
    newSheet[f"D{row_counters[category]}"] = grade
    
    # Increment row counter for this sheet
    row_counters[category] += 1

# Now add summary statistics to each sheet
for category, grades in category_grades.items():
    sheet = newWorkbook[category]
    
    # Calculate statistics
    highest_grade = max(grades)
    lowest_grade = min(grades)
    mean_grade = statistics.mean(grades)
    median_grade = statistics.median(grades)
    student_count = len(grades)
    
    # Add statistics in column F
    sheet["F1"] = "Summary Statistics"
    sheet["G1"] = "Value"
    sheet["F2"] = "Highest Grade"
    sheet["F3"] = "Lowest Grade"
    sheet["F4"] = "Mean Grade"
    sheet["F5"] = "Median Grade"
    sheet["F6"] = "Number of Students"
    
    
    # Add data in column G
    sheet["G2"] = highest_grade
    sheet["G3"] = lowest_grade
    sheet["G4"] = round(mean_grade, 2)  # Round to 2 decimal places
    sheet["G5"] = round(median_grade, 2)  # Round to 2 decimal places
    sheet["G6"] = student_count

    # Bold headers in columns A, B, C, D, F, G
    for col in ["A", "B", "C", "D", "F", "G"]:
        sheet[f"{col}1"].font = Font(bold=True)
    
    # Adjust column widths based on header length + 5
    column_widths = {
        "A": len("Last Name") + 5,
        "B": len("First Name") + 5,
        "C": len("Student ID") + 5,
        "D": len("Grade") + 5,
        "F": len("Highest Grade") + 5,
        "G": len("Number of Students") + 5
    }
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width
# Save the workbook
newWorkbook.save("Organized_Data.xlsx")

newWorkbook.close() # Closes the new workbook
myWorkbook.close() # Closes the loaded workbook
